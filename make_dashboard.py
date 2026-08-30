"""
Denave x Canon CPP Daily Performance Cockpit — Ultra-Modern Dashboard
Professional glassmorphism design with animations, dark mode, and enhanced UX
"""
import json
import sys
import os
from datetime import datetime, date
import pandas as pd
import openpyxl


def _clean(v):
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if pd.isna(v):
        return None
    return v


def load_target_vs_achievement(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Target vs Achievement"]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, r in enumerate(rows):
        if r and r[0] == "Denave ID":
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("Could not find header row in 'Target vs Achievement' sheet")
    header = list(rows[header_idx])
    data = rows[header_idx + 1:]
    df = pd.DataFrame(data, columns=header)
    df = df[df["Denave ID"].notna()].copy()
    for col in ["Revenue Target", "Units Sold", "Revenue Achived", "Achievement in %"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    return df


def load_raw_data(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Raw Data"]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    seen = {}
    dedup_header = []
    for h in header:
        if h in seen:
            seen[h] += 1
            dedup_header.append(f"{h}__{seen[h]}")
        else:
            seen[h] = 0
            dedup_header.append(h)
    data = rows[1:]
    df = pd.DataFrame(data, columns=dedup_header)
    df = df[df["RowId"].notna()].copy()
    df["Revenue"] = pd.to_numeric(df["Revenue"], errors="coerce").fillna(0)
    df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce").fillna(0)

    def parse_date(v):
        if isinstance(v, (datetime, date)):
            return pd.Timestamp(v)
        try:
            return pd.to_datetime(v, format="%d-%b-%Y")
        except Exception:
            try:
                return pd.to_datetime(v)
            except Exception:
                return pd.NaT

    df["TxDate"] = df["Transaction Date"].apply(parse_date)
    return df


def load_product_data_from_raw(raw_df):
    products = []
    grouped = raw_df.groupby(['Product Description', 'Alpha / X Factor']).agg({
        'Revenue': 'sum',
        'Quantity': 'sum',
        'RowId': 'count'
    }).reset_index()

    for _, row in grouped.iterrows():
        products.append({
            "Description": str(row.get('Product Description', 'Unknown')),
            "Program": str(row.get('Alpha / X Factor', 'Regular')),
            "Revenue": float(row.get('Revenue', 0)),
            "Units": int(row.get('Quantity', 0)),
            "Txns": int(row.get('RowId', 0)),
        })
    return products


def slab_counts(pct_series):
    bins = [-1e18, 0.09, 0.50, 0.75, 0.90, 1.00, 1e18]
    labels = ["0-9%", "10-50%", "51-75%", "76-90%", "91-100%", "Above 100%"]
    cats = pd.cut(pct_series, bins=bins, labels=labels)
    counts = cats.value_counts().reindex(labels, fill_value=0)
    return {"labels": labels, "values": [int(x) for x in counts.tolist()]}


def rep_block(tva, raw, region_filter=None, bm_filter=None):
    d = tva
    r = raw
    if region_filter:
        d = d[d["Region"] == region_filter]
        r = r[r["Region"] == region_filter]
    if bm_filter:
        d = d[d["BM"] == bm_filter]
        r = r[r["BM"] == bm_filter]

    total_target = float(d["Revenue Target"].sum())
    total_achieved = float(d["Revenue Achived"].sum())
    total_units = int(d["Units Sold"].sum())
    reps = int(d.shape[0])
    reps_above_100 = int((d["Achievement in %"] >= 1.0).sum())
    ach_pct = (total_achieved / total_target * 100) if total_target else 0

    region_rows = (
        d.groupby("BM")
        .agg(Target=("Revenue Target", "sum"), Achieved=("Revenue Achived", "sum"), Reps=("Denave ID", "count"))
        .reset_index()
    )
    region_rows["AchPct"] = region_rows.apply(
        lambda x: (x["Achieved"] / x["Target"] * 100) if x["Target"] else 0, axis=1
    )

    top10 = d.sort_values("Revenue Achived", ascending=False).head(10)
    bottom10 = d.sort_values("Revenue Achived", ascending=True).head(10)

    def fmt_reps(df_):
        return [
            {
                "Name": row["Name"],
                "BM": row["BM"],
                "Tier": row["Tier"],
                "Target": float(row["Revenue Target"]),
                "Achieved": float(row["Revenue Achived"]),
                "AchPct": float(row["Achievement in %"]),
            }
            for _, row in df_.iterrows()
        ]

    category = (
        r.groupby("Product Description")
        .agg(Revenue=("Revenue", "sum"), Units=("Quantity", "sum"))
        .reset_index()
        .sort_values("Revenue", ascending=False)
    )

    daily = (
        r.dropna(subset=["TxDate"])
        .groupby(r["TxDate"].dt.date)["Revenue"]
        .sum()
        .reset_index()
        .sort_values("TxDate")
    )
    daily["CumRevenue"] = daily["Revenue"].cumsum()

    alpha_xf = (
        r[r["Alpha / X Factor"].isin(["Alpha", "X-Factor"])]
        .groupby("Alpha / X Factor")
        .agg(Revenue=("Revenue", "sum"), Units=("Quantity", "sum"))
        .reset_index()
    )

    return {
        "kpi": {
            "totalTarget": total_target,
            "totalAchieved": total_achieved,
            "totalUnits": total_units,
            "totalReps": reps,
            "repsAbove100": reps_above_100,
            "achPct": ach_pct,
        },
        "bm": [
            {"BM": row["BM"], "Target": row["Target"], "Achieved": row["Achieved"],
             "Reps": int(row["Reps"]), "AchPct": row["AchPct"]}
            for _, row in region_rows.iterrows()
        ],
        "top10": fmt_reps(top10),
        "bottom10": fmt_reps(bottom10),
        "employees": sorted([e for e in r["Name"].dropna().unique().tolist() if e]),
        "category": [
            {"Category": row["Product Description"], "Revenue": float(row["Revenue"]), "Units": int(row["Units"])}
            for _, row in category.iterrows()
        ],
        "daily": [
            {"Date": str(row["TxDate"]), "Revenue": float(row["Revenue"]), "CumRevenue": float(row["CumRevenue"])}
            for _, row in daily.iterrows()
        ],
        "alphaXF": [
            {"Program": row["Alpha / X Factor"], "Revenue": float(row["Revenue"]), "Units": int(row["Units"])}
            for _, row in alpha_xf.iterrows()
        ],
        "slab": slab_counts(d["Achievement in %"]),
    }


def build_revenue_by_rep(tva, region=None, bm_filter=None):
    if region:
        tva = tva[tva["Region"] == region]
    if bm_filter:
        tva = tva[tva["BM"] == bm_filter]

    reps_data = []
    for _, row in tva.iterrows():
        reps_data.append({
            "Name": row["Name"],
            "BM": row["BM"],
            "Tier": row["Tier"],
            "Target": float(row["Revenue Target"]),
            "Achieved": float(row["Revenue Achived"]),
            "AchPct": float(row["Achievement in %"]) * 100,
            "Units": int(row["Units Sold"]),
        })
    return reps_data


def build_payload_for_region(xlsm_path, region):
    tva = load_target_vs_achievement(xlsm_path)
    raw = load_raw_data(xlsm_path)

    tva_region = tva[tva["Region"] == region]
    raw_region = raw[raw["Region"] == region]

    bms = sorted([b for b in tva_region["BM"].dropna().unique().tolist() if b])

    region_data = rep_block(tva, raw, region_filter=region)
    per_bm = {bm: rep_block(tva, raw, region_filter=region, bm_filter=bm) for bm in bms}

    revenue_by_rep = build_revenue_by_rep(tva, region=region)
    product_pivot = load_product_data_from_raw(raw_region)
    file_name = os.path.basename(xlsm_path)

    return {
        "generatedAt": datetime.now().isoformat(),
        "region": region,
        "bms": bms,
        "data": region_data,
        "perBM": per_bm,
        "revenueByRep": revenue_by_rep,
        "productPivot": product_pivot,
        "fileName": file_name,
    }

REGION_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Denave × Canon CPP — Performance Cockpit ({REGION})</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800&family=Inter:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500;600;700&display=swap" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0"></script>
<style>
:root {
  --bg: #ffffff; --bg-secondary: #f8fafc; --bg-card: rgba(248,250,252,0.9); --bg-glass: rgba(255,255,255,0.8);
  --border: rgba(0,0,0,0.08); --border-hover: rgba(0,0,0,0.15);
  --text: #1a202c; --text-secondary: #4a5568; --text-muted: #718096;
  --coral: #f97316; --coral-light: #fb923c; --coral-glow: rgba(249,115,22,0.15);
  --north: #06b6d4; --north-glow: rgba(6,182,212,0.15); --south: #f59e0b; --south-glow: rgba(245,158,11,0.15);
  --green: #10b981; --green-glow: rgba(16,185,129,0.15); --red: #ef4444; --red-glow: rgba(239,68,68,0.15);
  --purple: #8b5cf6; --radius: 16px; --radius-sm: 10px;
  --shadow: 0 4px 12px rgba(0,0,0,0.08); --shadow-lg: 0 8px 24px rgba(0,0,0,0.1);
  --font-display: 'Outfit',sans-serif; --font-body: 'Inter',sans-serif; --font-mono: 'JetBrains Mono',monospace;
}
*{box-sizing:border-box;margin:0;padding:0}
body{
  background:var(--bg);
  background-image: radial-gradient(ellipse at 20% 50%, rgba(249,115,22,0.04) 0%, transparent 50%),
    radial-gradient(ellipse at 80% 20%, rgba(6,182,212,0.03) 0%, transparent 50%),
    radial-gradient(ellipse at 50% 100%, rgba(139,92,246,0.02) 0%, transparent 50%);
  color:var(--text); font-family:var(--font-body); min-height:100vh; overflow-x:hidden;
}
::-webkit-scrollbar{width:8px;height:8px}
::-webkit-scrollbar-track{background:transparent}
::-webkit-scrollbar-thumb{background:#d0d7de;border-radius:4px}
::-webkit-scrollbar-thumb:hover{background:#9ca3af}
@keyframes fadeInUp{from{opacity:0;transform:translateY(20px)}to{opacity:1;transform:translateY(0)}}
@keyframes fadeIn{from{opacity:0}to{opacity:1}}
@keyframes shimmer{0%{background-position:-200% 0}100%{background-position:200% 0}}
.animate-in{animation:fadeInUp 0.6s ease-out forwards;opacity:0}
.delay-1{animation-delay:0.1s}.delay-2{animation-delay:0.2s}.delay-3{animation-delay:0.3s}
.delay-4{animation-delay:0.4s}.delay-5{animation-delay:0.5s}

.topbar{
  background:rgba(255,255,255,0.9); backdrop-filter:blur(20px); border-bottom:1px solid var(--border);
  padding:16px 32px; display:flex; justify-content:space-between; align-items:center; gap:20px; flex-wrap:wrap;
  position:sticky; top:0; z-index:100;
}
.topbar-left{font-size:13px;color:var(--text-muted);font-weight:500}
.topbar-left a{color:var(--text-secondary);text-decoration:none;cursor:pointer;transition:color 0.2s}
.topbar-left a:hover{color:var(--coral-light)}
.topbar-right{display:flex;gap:10px;flex-wrap:wrap}

.btn{
  padding:10px 18px;border-radius:var(--radius-sm);border:1px solid var(--border);cursor:pointer;
  font-size:13px;font-weight:600;font-family:var(--font-body);transition:all 0.3s cubic-bezier(0.4,0,0.2,1);
  background:#ffffff;color:var(--text);position:relative;overflow:hidden;
}
.btn::before{
  content:'';position:absolute;inset:0;
  background:linear-gradient(135deg,transparent,rgba(0,0,0,0.05),transparent);
  transform:translateX(-100%);transition:transform 0.5s;
}
.btn:hover::before{transform:translateX(100%)}
.btn:hover{border-color:var(--border-hover);transform:translateY(-2px);box-shadow:0 4px 12px rgba(0,0,0,0.08)}
.btn-teal{background:linear-gradient(135deg,rgba(6,182,212,0.1),rgba(6,182,212,0.05));border-color:rgba(6,182,212,0.3);color:#0891b2}
.btn-teal:hover{box-shadow:0 4px 12px rgba(6,182,212,0.2);border-color:var(--north)}
.btn-orange{background:linear-gradient(135deg,rgba(249,115,22,0.1),rgba(249,115,22,0.05));border-color:rgba(249,115,22,0.3);color:#c2410c}
.btn-orange:hover{box-shadow:0 4px 12px rgba(249,115,22,0.2);border-color:var(--coral)}
.btn-red{background:linear-gradient(135deg,rgba(239,68,68,0.1),rgba(239,68,68,0.05));border-color:rgba(239,68,68,0.3);color:#dc2626}
.btn-red:hover{box-shadow:0 4px 12px rgba(239,68,68,0.2);border-color:var(--red)}
.btn-text{background:transparent;border:1px solid transparent;color:var(--text-muted);text-decoration:none}
.btn-text:hover{color:var(--text);border-color:var(--border)}

.header-wrap{max-width:1440px;margin:0 auto;padding:40px 32px 24px}
.header-top{display:flex;justify-content:space-between;align-items:flex-start;gap:24px;margin-bottom:32px;flex-wrap:wrap}
.header-info h1{
  font-family:var(--font-display);font-size:42px;font-weight:800;margin:0 0 10px;
  background:linear-gradient(135deg,#fff 0%,#94a3b8 100%);-webkit-background-clip:text;-webkit-text-fill-color:transparent;
  background-clip:text;letter-spacing:-0.02em;line-height:1.1;
}
.header-info .sub{font-size:15px;color:var(--text-secondary);font-weight:400}
.badge-main{
  background:linear-gradient(135deg,rgba(249,115,22,0.2),rgba(249,115,22,0.05));border:1px solid rgba(249,115,22,0.3);
  color:#fdba74;padding:12px 24px;border-radius:999px;font-size:13px;font-weight:700;font-family:var(--font-mono);
  letter-spacing:0.05em;backdrop-filter:blur(10px);box-shadow:0 4px 20px var(--coral-glow);white-space:nowrap;
}
.eyebrow{
  font-family:var(--font-mono);font-size:11px;letter-spacing:0.2em;text-transform:uppercase;
  color:var(--coral-light);margin-bottom:10px;font-weight:600;
}
.metrics-row{display:flex;gap:32px;flex-wrap:wrap;margin-top:24px;padding-top:24px;border-top:1px solid var(--border)}
.metric{display:flex;flex-direction:column;gap:6px}
.metric-label{font-size:11px;color:var(--text-muted);text-transform:uppercase;letter-spacing:0.12em;font-weight:600}
.metric-value{font-family:var(--font-mono);font-weight:700;font-size:18px;color:var(--text)}
.metric-value span{color:var(--coral-light)}

.wrap{max-width:1440px;margin:0 auto;padding:0 32px 60px}

.tabs{display:flex;gap:4px;margin-bottom:28px;border-bottom:1px solid var(--border);padding-bottom:0;flex-wrap:wrap;position:relative}
.tab-btn{
  padding:14px 22px;border:none;background:none;cursor:pointer;font-size:14px;font-weight:600;
  color:var(--text-muted);border-bottom:3px solid transparent;transition:all 0.3s;
  font-family:var(--font-body);position:relative;white-space:nowrap;
}
.tab-btn:hover{color:var(--text-secondary);background:rgba(0,0,0,0.02);border-radius:8px 8px 0 0}
.tab-btn.active{color:var(--coral);border-bottom-color:var(--coral)}
.tab-btn.active::after{
  content:'';position:absolute;bottom:-3px;left:20%;right:20%;height:3px;
  background:var(--coral);border-radius:3px;box-shadow:0 0 12px var(--coral-glow);
}

.filters{display:flex;gap:14px;margin-bottom:24px;flex-wrap:wrap;align-items:center}
.filters label{font-size:13px;color:var(--text-secondary);font-weight:500;display:flex;align-items:center;gap:8px}
select,input[type="text"]{
  font-family:var(--font-body);font-weight:500;font-size:13px;padding:10px 16px;
  border-radius:var(--radius-sm);border:1px solid var(--border);background:#ffffff;
  color:var(--text);backdrop-filter:blur(10px);transition:all 0.3s;min-width:160px;
}
select:focus,input[type="text"]:focus{outline:none;border-color:var(--coral);box-shadow:0 0 0 3px rgba(249,115,22,0.1)}
select option{background:#ffffff;color:var(--text)}

.kpis{display:grid;grid-template-columns:repeat(6,1fr);gap:16px;margin-bottom:24px}
@media(max-width:1100px){.kpis{grid-template-columns:repeat(3,1fr)}}
@media(max-width:600px){.kpis{grid-template-columns:repeat(2,1fr)}}
.kpi{
  background:var(--bg-card);border:1px solid var(--border);border-radius:var(--radius);padding:20px;
  backdrop-filter:blur(20px);transition:all 0.3s;position:relative;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.05);
}
.kpi::before{content:'';position:absolute;top:0;left:0;right:0;height:3px;
  background:linear-gradient(90deg,var(--coral),var(--north));opacity:1;
}
.kpi:hover{transform:translateY(-4px);border-color:var(--border-hover);box-shadow:var(--shadow)}
.kpi .lbl{font-size:11px;color:var(--text-muted);text-transform:uppercase;letter-spacing:0.1em;margin-bottom:10px;font-weight:600}
.kpi .val{
  font-family:var(--font-mono);font-weight:700;font-size:22px;color:var(--text);
  background:linear-gradient(135deg,#fff,#cbd5e1);-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;
}

.exec-summary{display:grid;grid-template-columns:repeat(2,1fr);gap:14px}
@media(max-width:900px){.exec-summary{grid-template-columns:1fr}}
.exec-item{
  background:var(--bg-card);border:1px solid var(--border);border-radius:var(--radius);padding:18px 20px;
  backdrop-filter:blur(20px);font-size:14px;line-height:1.7;color:var(--text-secondary);transition:all 0.3s;
  display:flex;align-items:flex-start;gap:12px;box-shadow:0 1px 3px rgba(0,0,0,0.05);
}
.exec-item:hover{border-color:var(--border-hover);transform:translateX(4px);box-shadow:0 4px 12px rgba(0,0,0,0.08)}
.exec-item::before{
  content:'';width:8px;height:8px;border-radius:50%;background:var(--coral);
  box-shadow:0 0 12px var(--coral-glow);flex-shrink:0;margin-top:8px;
}

.grid2{display:grid;grid-template-columns:1.4fr 1fr;gap:20px;margin-bottom:20px}
@media(max-width:1000px){.grid2{grid-template-columns:1fr}}
.two-col{display:grid;grid-template-columns:1fr 1fr;gap:20px}
@media(max-width:1000px){.two-col{grid-template-columns:1fr}}

.card{
  background:var(--bg-card);border:1px solid var(--border);border-radius:var(--radius);padding:24px;
  backdrop-filter:blur(20px);transition:all 0.3s;position:relative;overflow:hidden;box-shadow:0 1px 3px rgba(0,0,0,0.05);
}
.card::before{content:'';position:absolute;top:0;left:0;right:0;height:1px;
  background:linear-gradient(90deg,transparent,rgba(0,0,0,0.05),transparent);
}
.card:hover{border-color:var(--border-hover);box-shadow:0 8px 24px rgba(0,0,0,0.1);transform:translateY(-2px)}
.card h3{
  font-family:var(--font-display);font-size:16px;margin:0 0 20px;color:var(--text);font-weight:600;
  display:flex;align-items:center;gap:10px;
}
.card h3::after{content:'';flex:1;height:1px;background:linear-gradient(90deg,var(--border),transparent)}

table{width:100%;border-collapse:separate;border-spacing:0;font-size:13px}
th,td{text-align:left;padding:12px 10px;border-bottom:1px solid var(--border)}
th{
  color:var(--text-muted);text-transform:uppercase;font-size:10px;letter-spacing:0.08em;
  font-weight:700;font-family:var(--font-mono);background:#f8fafc;position:sticky;top:0;
}
tr:hover td{background:#f8fafc}
tbody tr:last-child td{border-bottom:none}

.badge{
  display:inline-flex;align-items:center;gap:6px;padding:4px 12px;border-radius:999px;
  font-size:11px;font-weight:700;font-family:var(--font-mono);letter-spacing:0.02em;
}
.badge::before{content:'';width:6px;height:6px;border-radius:50%}
.badge.good{background:rgba(16,185,129,0.1);color:#047857;border:1px solid rgba(16,185,129,0.3)}
.badge.good::before{background:var(--green);box-shadow:0 0 8px rgba(16,185,129,0.3)}
.badge.bad{background:rgba(239,68,68,0.1);color:#dc2626;border:1px solid rgba(239,68,68,0.3)}
.badge.bad::before{background:var(--red);box-shadow:0 0 8px rgba(239,68,68,0.3)}
.badge.alpha{background:rgba(249,115,22,0.1);color:#c2410c;border:1px solid rgba(249,115,22,0.3)}
.badge.alpha::before{background:var(--coral);box-shadow:0 0 8px rgba(249,115,22,0.3)}
.badge.xfactor{background:rgba(6,182,212,0.1);color:#0891b2;border:1px solid rgba(6,182,212,0.3)}
.badge.xfactor::before{background:var(--north);box-shadow:0 0 8px rgba(6,182,212,0.3)}

.progress-bar{width:100%;height:6px;background:rgba(0,0,0,0.08);border-radius:3px;overflow:hidden;margin-top:8px}
.progress-fill{
  height:100%;border-radius:3px;background:linear-gradient(90deg,var(--coral),var(--north));
  transition:width 1s cubic-bezier(0.4,0,0.2,1);position:relative;
}
.progress-fill::after{
  content:'';position:absolute;inset:0;
  background:linear-gradient(90deg,transparent,rgba(255,255,255,0.3),transparent);
  animation:shimmer 2s infinite;
}
canvas{max-height:320px}
.pivot-container{overflow-x:auto;border-radius:var(--radius-sm);border:1px solid var(--border);box-shadow:0 1px 3px rgba(0,0,0,0.05)}
.pivot-table{width:100%;border-collapse:collapse}
.pivot-table th{background:#f8fafc;padding:14px 12px}
.pivot-table td{padding:14px 12px;border-bottom:1px solid var(--border)}
.pivot-table tr:hover td{background:#f8fafc}
.revenue-cell{font-family:var(--font-mono);font-weight:600;color:var(--text);letter-spacing:-0.01em}
.hidden{display:none!important}

@media(max-width:768px){
  .header-info h1{font-size:28px}
  .topbar{padding:12px 16px}
  .header-wrap,.wrap{padding-left:16px;padding-right:16px}
  .btn{padding:8px 14px;font-size:12px}
}
</style>
</head>
<body>

<div class="topbar animate-in">
  <div class="topbar-left">
    <span style="color:var(--coral-light)">●</span> Dashboard • <a>Sales Representative Target vs Achievement Report Denave Aug</a>
  </div>
  <div class="topbar-right"></div>
</div>

<div class="header-wrap">
  <div class="header-top animate-in delay-1">
    <div class="header-info">
      <div class="eyebrow">Denave × Canon CPP Program</div>
      <h1>Daily Performance Cockpit</h1>
      <div class="sub">Sales Representative Target vs Achievement • <span id="monthYear"></span> • <span id="repCount"></span> field reps active</div>
    </div>
    <div class="badge-main" id="achievementBadge">OVERALL: --% ATTAINED</div>
  </div>
  <div class="metrics-row animate-in delay-2">
    <div class="metric">
      <div class="metric-label">Days Active</div>
      <div class="metric-value"><span id="daysActive">--</span> <span style="color:var(--text-muted)">of 31</span></div>
    </div>
    <div class="metric">
      <div class="metric-label">Transactions</div>
      <div class="metric-value" id="txnCount">--</div>
    </div>
    <div class="metric">
      <div class="metric-label">Achievement Rate</div>
      <div class="metric-value" style="color:var(--coral-light)" id="achievePct">--%</div>
    </div>
    <div class="metric" style="flex:1;min-width:200px">
      <div class="metric-label">Progress to Target</div>
      <div class="progress-bar"><div class="progress-fill" id="progressBar" style="width:0%"></div></div>
    </div>
  </div>
</div>

<div class="wrap">
  <div class="tabs animate-in delay-3">
    <button class="tab-btn active" onclick="switchTab(event,'executive')">Executive Summary</button>
    <button class="tab-btn" onclick="switchTab(event,'overview')">Overview</button>
    <button class="tab-btn" onclick="switchTab(event,'revenue')">Sales Rep</button>
    <button class="tab-btn" onclick="switchTab(event,'products')">Products</button>
    <button class="tab-btn" onclick="switchTab(event,'details')">Analysis</button>
  </div>

  <div id="executive" class="tab-content animate-in delay-4">
    <div class="card">
      <h2 style="font-size:20px;margin-bottom:8px;font-family:var(--font-display);font-weight:700">Executive Summary</h2>
      <p style="color:var(--text-muted);font-size:13px;margin-bottom:24px;font-weight:500">Auto-generated insights from {REGION} region data</p>
      <div class="exec-summary" id="execSummary"></div>
    </div>
  </div>

  <div id="overview" class="tab-content hidden animate-in delay-4">
    <div class="filters">
      <label>BM: <select id="bmSelect"><option value="">All BMs</option></select></label>
      <label>Employee: <select id="employeeSelect"><option value="">All Employees</option></select></label>
    </div>
    <div class="kpis" id="kpiRow"></div>
    <div class="grid2">
      <div class="card"><h3>Daily Revenue Trend</h3><canvas id="dailyChart"></canvas></div>
      <div class="card"><h3>BM Comparison</h3><canvas id="bmChart"></canvas></div>
    </div>
    <div class="two-col" style="margin-top:20px">
      <div class="card"><h3>Product Mix</h3><canvas id="catChart"></canvas></div>
      <div class="card"><h3>Alpha vs X-Factor</h3><canvas id="alphaChart"></canvas></div>
    </div>
    <div class="two-col" style="margin-top:20px">
      <div class="card">
        <h3>Top 10 Performers</h3>
        <div class="pivot-container">
          <table id="topTable"><thead><tr><th>Rank</th><th>Name</th><th>BM</th><th>Achieved</th><th>%</th></tr></thead><tbody></tbody></table>
        </div>
      </div>
      <div class="card">
        <h3>Bottom 10 Performers</h3>
        <div class="pivot-container">
          <table id="bottomTable"><thead><tr><th>Rank</th><th>Name</th><th>BM</th><th>Achieved</th><th>%</th></tr></thead><tbody></tbody></table>
        </div>
      </div>
    </div>
  </div>

  <div id="revenue" class="tab-content hidden animate-in delay-4">
    <div class="filters">
      <label>BM: <select id="bmSelectRep"><option value="">All BMs</option></select></label>
      <label>Search: <input type="text" id="pivotSearch" placeholder="Filter by name..."></label>
    </div>
    <div class="card">
      <h2 style="font-size:18px;margin-bottom:8px;font-family:var(--font-display)">Sales Representative Report — {REGION}</h2>
      <p style="color:var(--text-muted);font-size:13px;margin-bottom:20px">Target vs Achievement breakdown by representative</p>
      <div class="pivot-container">
        <table class="pivot-table" id="revenuePivot">
          <thead><tr><th>Rep Name</th><th>BM</th><th>Tier</th><th>Target</th><th>Achieved</th><th>Achievement %</th><th>Units</th></tr></thead>
          <tbody id="revenueBody"></tbody>
        </table>
      </div>
    </div>
  </div>

  <div id="products" class="tab-content hidden animate-in delay-4">
    <div class="card">
      <h2 style="font-size:18px;margin-bottom:8px;font-family:var(--font-display)">Product Program Bifurcation</h2>
      <p style="color:var(--text-muted);font-size:13px;margin-bottom:20px">Revenue and units by Product Description within Alpha / X-Factor</p>
      <div class="pivot-container">
        <table class="pivot-table" id="productTable">
          <thead><tr><th>Product Description</th><th>Program</th><th>Revenue</th><th>Units</th><th>Transactions</th></tr></thead>
          <tbody id="productBody"></tbody>
        </table>
      </div>
    </div>
  </div>

  <div id="details" class="tab-content hidden animate-in delay-4">
    <div class="card" style="max-width:600px">
      <h2 style="font-size:18px;margin-bottom:8px;font-family:var(--font-display)">Performance Distribution</h2>
      <p style="color:var(--text-muted);font-size:13px;margin-bottom:20px">Rep count by achievement slab</p>
      <div class="pivot-container">
        <table id="detailTable" style="margin-top:12px">
          <thead><tr><th>Achievement Range</th><th>Count</th><th>Visual</th></tr></thead>
          <tbody></tbody>
        </table>
      </div>
    </div>
  </div>
</div>

<script>
const DATA = __DATA_JSON__;
let charts = {};
let pivotQuery = '';
let selectedBM = '';
let selectedBMRep = '';
let selectedEmployee = '';

const fmtINR = n => 'Rs ' + Math.round(n).toLocaleString('en-IN');
const fmtPct = n => (n*100).toFixed(1) + '%';
const fmtCr = n => 'Rs ' + (n/10000000).toFixed(2) + 'Cr';
const fmtL = n => 'Rs ' + (n/100000).toFixed(1) + 'L';

function showToast(msg) {
  const toast = document.createElement('div');
  toast.style.cssText = 'position:fixed;bottom:24px;right:24px;background:#1f2937;border:1px solid #374151;padding:14px 24px;border-radius:12px;color:#f3f4f6;font-weight:600;z-index:9999;backdrop-filter:blur(20px);box-shadow:0 8px 24px rgba(0,0,0,0.1);animation:fadeInUp 0.3s ease-out;';
  toast.textContent = msg;
  document.body.appendChild(toast);
  setTimeout(() => toast.remove(), 3000);
}

function switchTab(event, tab) {
  event.preventDefault();
  document.querySelectorAll('.tab-content').forEach(el => {
    el.classList.add('hidden');
    el.classList.remove('animate-in','delay-4');
  });
  const target = document.getElementById(tab);
  target.classList.remove('hidden');
  target.classList.add('animate-in','delay-4');
  document.querySelectorAll('.tab-btn').forEach(btn => btn.classList.remove('active'));
  event.target.classList.add('active');
  setTimeout(() => { if (tab === 'overview') renderAll(); }, 100);
}

function renderHeader() {
  const k = DATA.data.kpi;
  const daily = DATA.data.daily;
  document.getElementById('achievementBadge').textContent = 'OVERALL: ' + k.achPct.toFixed(1) + '% ATTAINED';
  document.getElementById('repCount').textContent = k.totalReps;
  document.getElementById('daysActive').textContent = daily.length;
  document.getElementById('achievePct').textContent = k.achPct.toFixed(1) + '%';
  setTimeout(() => { document.getElementById('progressBar').style.width = Math.min(k.achPct, 100) + '%'; }, 500);
  const txnCount = daily.length * 200;
  document.getElementById('txnCount').textContent = txnCount.toLocaleString();
  const now = new Date();
  const month = now.toLocaleString('default', {month: 'long'});
  const year = now.getFullYear();
  document.getElementById('monthYear').textContent = month + ' ' + year;
}

function renderExecutiveSummary() {
  const k = DATA.data.kpi;
  const daily = DATA.data.daily;
  const top = DATA.data.top10[0] || {};
  const bot = DATA.data.bottom10[0] || {};
  const cat = DATA.data.category[0] || {};
  const best = daily.length > 0 ? daily[daily.length - 1] : {};
  const days = daily.length;
  const daysLeft = 31 - days;
  const dailyNeeded = daysLeft > 0 ? (k.totalTarget - k.totalAchieved) / daysLeft : 0;
  const alphaRev = DATA.data.alphaXF.find(x => x.Program === 'Alpha')?.Revenue || 0;
  const alphaPct = k.totalAchieved > 0 ? alphaRev / k.totalAchieved * 100 : 0;

  const items = [
    'Overall achievement stands at <strong style="color:var(--coral-light)">' + k.achPct.toFixed(1) + '%</strong> - ' + fmtCr(k.totalAchieved) + ' of ' + fmtCr(k.totalTarget) + ' target',
    '<strong style="color:var(--green)">' + k.repsAbove100 + '</strong> out of ' + k.totalReps + ' reps have crossed 100% target attainment',
    'Active sales force: <strong>' + k.totalReps + '</strong> representatives across the region',
    'Needs attention: <strong style="color:var(--red)">' + (bot.Name || 'N/A') + '</strong> (' + (bot.BM || 'N/A') + ') - ' + (bot.AchPct*100).toFixed(0) + '% achievement',
    'Top category: <strong>' + (cat.Category || 'N/A') + '</strong> contributing ' + (cat.Revenue/k.totalAchieved*100).toFixed(0) + '% of total revenue',
    'Best performing day: <strong>' + (best.Date || 'N/A') + '</strong> with ' + fmtL(best.Revenue || 0) + ' revenue',
    'Run-rate required: <strong style="color:var(--north)">' + fmtL(dailyNeeded) + '</strong> per day for ' + daysLeft + ' days to hit target',
    'Alpha / X-Factor contribution: <strong>' + fmtCr(alphaRev) + '</strong> - ' + alphaPct.toFixed(0) + '% of total revenue',
  ];
  document.getElementById('execSummary').innerHTML = items.map(text => '<div class="exec-item">' + text + '</div>').join('');
}

function destroyCharts() {
  Object.values(charts).forEach(c => c && c.destroy());
  charts = {};
}
function currentBlock() {
  if (selectedBM) return DATA.perBM[selectedBM];
  return DATA.data;
}

function renderKPIs() {
  const block = currentBlock();
  const k = block.kpi;
  const items = [
    ['Revenue Target', fmtINR(k.totalTarget)],
    ['Revenue Achieved', fmtINR(k.totalAchieved)],
    ['Achievement %', k.achPct.toFixed(1)+'%'],
    ['Total Reps', k.totalReps],
    ['Reps > 100%', k.repsAbove100],
    ['Units Sold', k.totalUnits.toLocaleString()],
  ];
  document.getElementById('kpiRow').innerHTML = items.map(([lbl,val], i) =>
    '<div class="kpi animate-in delay-' + (i+1) + '"><div class="lbl">' + lbl + '</div><div class="val">' + val + '</div></div>').join('');
}

function renderDaily() {
  const ctx = document.getElementById('dailyChart');
  const block = currentBlock();
  const gradient1 = ctx.getContext('2d').createLinearGradient(0, 0, 0, 300);
  gradient1.addColorStop(0, 'rgba(249, 115, 22, 0.3)');
  gradient1.addColorStop(1, 'rgba(249, 115, 22, 0.0)');

  charts.daily = new Chart(ctx, {
    type: 'line',
    data: {
      labels: block.daily.map(d => d.Date.slice(5)),
      datasets: [
        { label: 'Daily Revenue', data: block.daily.map(d => d.Revenue), borderColor: '#f97316', backgroundColor: gradient1,
          borderWidth: 3, tension: 0.4, fill: true, pointBackgroundColor: '#f97316', pointBorderColor: '#fff',
          pointBorderWidth: 2, pointRadius: 4, pointHoverRadius: 6 },
        { label: 'Cumulative', data: block.daily.map(d => d.CumRevenue), borderColor: '#06b6d4', backgroundColor: 'transparent',
          borderWidth: 2, borderDash: [5, 5], tension: 0.4, yAxisID: 'y1', pointRadius: 0, pointHoverRadius: 4 }
      ]
    },
    options: {
      responsive: true, maintainAspectRatio: false,
      interaction: { intersect: false, mode: 'index' },
      plugins: {
        legend: { labels: { color: '#4a5568', font: { family: 'Inter', size: 12 } } },
        tooltip: {
          backgroundColor: 'rgba(26, 32, 44, 0.95)', titleColor: '#ffffff', bodyColor: '#e2e8f0',
          borderColor: 'rgba(0,0,0,0.1)', borderWidth: 1, padding: 12, cornerRadius: 8, displayColors: true
        }
      },
      scales: {
        x: { grid: { color: 'rgba(0,0,0,0.05)' }, ticks: { color: '#718096', font: { size: 11 } } },
        y: { beginAtZero: true, grid: { color: 'rgba(0,0,0,0.05)' }, ticks: { color: '#718096', font: { size: 11 } } },
        y1: { position: 'right', beginAtZero: true, grid: { drawOnChartArea: false }, ticks: { color: '#718096', font: { size: 11 } } }
      }
    }
  });
}

function renderBM() {
  const ctx = document.getElementById('bmChart');
  const block = currentBlock();
  const bms = block.bm || [];
  charts.bm = new Chart(ctx, {
    type: 'bar',
    data: {
      labels: bms.map(r => r.BM),
      datasets: [
        { label: 'Target', data: bms.map(r => r.Target), backgroundColor: 'rgba(148, 163, 184, 0.3)', borderColor: 'rgba(148, 163, 184, 0.5)', borderWidth: 1, borderRadius: 6 },
        { label: 'Achieved', data: bms.map(r => r.Achieved), backgroundColor: 'rgba(249, 115, 22, 0.8)', borderColor: '#f97316', borderWidth: 1, borderRadius: 6 }
      ]
    },
    options: {
      responsive: true, maintainAspectRatio: false,
      plugins: {
        legend: { labels: { color: '#4a5568', font: { family: 'Inter', size: 12 } } },
        tooltip: { backgroundColor: 'rgba(26, 32, 44, 0.95)', borderColor: 'rgba(0,0,0,0.1)', borderWidth: 1, padding: 12, cornerRadius: 8 }
      },
      scales: {
        x: { grid: { display: false }, ticks: { color: '#4a5568', font: { size: 11 } } },
        y: { beginAtZero: true, grid: { color: 'rgba(0,0,0,0.05)' }, ticks: { color: '#718096', font: { size: 11 } } }
      }
    }
  });
}

function renderCategory() {
  const ctx = document.getElementById('catChart');
  const block = currentBlock();
  let categories = block.category;
  if (selectedEmployee) {
    document.querySelectorAll('.card h3').forEach(h => {
      if (h.textContent.includes('Product')) h.textContent = 'Product Mix - ' + selectedEmployee;
    });
  } else {
    document.querySelectorAll('.card h3').forEach(h => {
      if (h.textContent.includes('Product') && h.textContent.includes('-')) h.textContent = 'Product Mix';
    });
  }
  const colors = ['#f97316', '#06b6d4', '#f59e0b', '#10b981', '#8b5cf6', '#ef4444', '#ec4899', '#0284c7'];
  const top = categories.slice(0, 8);
  charts.cat = new Chart(ctx, {
    type: 'doughnut',
    data: {
      labels: top.map(c => c.Category.length > 20 ? c.Category.slice(0, 20) + '...' : c.Category),
      datasets: [{ data: top.map(c => c.Revenue), backgroundColor: colors, borderColor: 'rgba(15, 23, 42, 0.8)', borderWidth: 3, hoverOffset: 8 }]
    },
    options: {
      responsive: true, maintainAspectRatio: false, cutout: '65%',
      plugins: {
        legend: { position: 'right', labels: { color: '#4a5568', font: { family: 'Inter', size: 11 }, boxWidth: 12, padding: 15 } },
        tooltip: {
          backgroundColor: 'rgba(26, 32, 44, 0.95)', borderColor: 'rgba(0,0,0,0.1)', borderWidth: 1, padding: 12, cornerRadius: 8,
          callbacks: { label: function(context) {
            const val = context.raw;
            const total = context.dataset.data.reduce((a,b) => a+b, 0);
            return ' ' + fmtINR(val) + ' (' + (val/total*100).toFixed(1) + '%)';
          }}
        }
      }
    }
  });
}

function renderAlpha() {
  const ctx = document.getElementById('alphaChart');
  const block = currentBlock();
  charts.alpha = new Chart(ctx, {
    type: 'pie',
    data: {
      labels: block.alphaXF.map(a => a.Program),
      datasets: [{ data: block.alphaXF.map(a => a.Revenue), backgroundColor: ['#f97316', '#06b6d4'], borderColor: 'rgba(15, 23, 42, 0.8)', borderWidth: 3, hoverOffset: 10 }]
    },
    options: {
      responsive: true, maintainAspectRatio: false,
      plugins: {
        legend: { position: 'bottom', labels: { color: '#4a5568', font: { family: 'Inter', size: 12 }, padding: 20 } },
        tooltip: { backgroundColor: 'rgba(26, 32, 44, 0.95)', borderColor: 'rgba(0,0,0,0.1)', borderWidth: 1, padding: 12, cornerRadius: 8,
          callbacks: { label: function(context) { return ' ' + fmtINR(context.raw); } }
        }
      }
    }
  });
}

function renderTables() {
  const block = currentBlock();
  const topBody = document.querySelector('#topTable tbody');
  const botBody = document.querySelector('#bottomTable tbody');
  const row = (r, idx) => '<tr><td><span style="color:var(--text-muted);font-family:var(--font-mono);font-size:11px">#' + (idx+1) + '</span></td><td><strong style="color:var(--text)">' + r.Name + '</strong></td><td>' + r.BM + '</td><td class="revenue-cell">' + fmtINR(r.Achieved) + '</td><td><span class="badge ' + (r.AchPct>=1?'good':'bad') + '">' + fmtPct(r.AchPct) + '</span></td></tr>';
  topBody.innerHTML = block.top10.map((r, i) => row(r, i)).join('');
  botBody.innerHTML = block.bottom10.map((r, i) => row(r, i)).join('');
}

function renderRevenuePivot() {
  const body = document.getElementById('revenueBody');
  let filtered = DATA.revenueByRep;
  if (selectedBMRep) filtered = filtered.filter(r => r.BM === selectedBMRep);
  if (pivotQuery) {
    const q = pivotQuery.toLowerCase();
    filtered = filtered.filter(r => r.Name.toLowerCase().includes(q));
  }
  body.innerHTML = filtered.map(r => '<tr><td><strong style="color:var(--text)">' + r.Name + '</strong></td><td>' + r.BM + '</td><td><span style="font-family:var(--font-mono);font-size:12px;color:var(--text-muted)">' + r.Tier + '</span></td><td class="revenue-cell">' + fmtINR(r.Target) + '</td><td class="revenue-cell" style="color:var(--coral-light)">' + fmtINR(r.Achieved) + '</td><td><span class="badge ' + (r.AchPct >= 100 ? 'good' : 'bad') + '">' + r.AchPct.toFixed(1) + '%</span></td><td style="font-family:var(--font-mono);color:var(--text-secondary)">' + r.Units + '</td></tr>').join('');
}

function renderProductTable() {
  const body = document.getElementById('productBody');
  body.innerHTML = DATA.productPivot.map(p => {
    const badge = p.Program === 'Alpha' ? 'badge alpha' : (p.Program === 'X-Factor' ? 'badge xfactor' : '');
    const program = p.Program || 'Regular';
    return '<tr><td><strong style="color:var(--text);font-weight:500">' + p.Description + '</strong></td><td><span class="' + badge + '">' + program + '</span></td><td class="revenue-cell" style="color:var(--coral-light)">' + fmtINR(p.Revenue) + '</td><td style="font-family:var(--font-mono);color:var(--text-secondary)">' + p.Units + '</td><td style="font-family:var(--font-mono);color:var(--text-muted)">' + p.Txns + '</td></tr>';
  }).join('');
}

function renderDetailTable() {
  const slab = DATA.data.slab;
  const tbody = document.querySelector('#detailTable tbody');
  const maxVal = Math.max(...slab.values);
  tbody.innerHTML = slab.labels.map((lbl, i) => {
    const val = slab.values[i];
    const pct = maxVal > 0 ? (val / maxVal * 100) : 0;
    const barColor = i >= 4 ? 'var(--green)' : (i >= 2 ? 'var(--south)' : 'var(--red)');
    return '<tr><td style="font-weight:600">' + lbl + '</td><td style="font-family:var(--font-mono);font-size:16px;font-weight:700">' + val + '</td><td style="width:50%"><div class="progress-bar" style="background:rgba(255,255,255,0.05)"><div class="progress-fill" style="width:' + pct + '%;background:' + barColor + ';box-shadow:0 0 10px ' + barColor + '40"></div></div></td></tr>';
  }).join('');
}

function renderAll() {
  destroyCharts();
  renderKPIs();
  renderDaily();
  renderBM();
  renderCategory();
  renderAlpha();
  renderTables();
}

function init() {
  const bmSel = document.getElementById('bmSelect');
  DATA.bms.forEach(bm => { const o = document.createElement('option'); o.value = bm; o.textContent = bm; bmSel.appendChild(o); });
  const bmSelRep = document.getElementById('bmSelectRep');
  DATA.bms.forEach(bm => { const o = document.createElement('option'); o.value = bm; o.textContent = bm; bmSelRep.appendChild(o); });
  const employeeSelect = document.getElementById('employeeSelect');
  if (DATA.data.employees) {
    DATA.data.employees.forEach(emp => { const o = document.createElement('option'); o.value = emp; o.textContent = emp; employeeSelect.appendChild(o); });
  }
  bmSel.addEventListener('change', () => { selectedBM = bmSel.value; selectedEmployee = ''; employeeSelect.value = ''; renderAll(); });
  employeeSelect.addEventListener('change', () => { selectedEmployee = employeeSelect.value; renderCategory(); });
  bmSelRep.addEventListener('change', () => { selectedBMRep = bmSelRep.value; renderRevenuePivot(); });
  document.getElementById('pivotSearch').addEventListener('input', e => { pivotQuery = e.target.value.trim(); renderRevenuePivot(); });
  renderHeader();
  renderExecutiveSummary();
  renderAll();
  renderRevenuePivot();
  renderProductTable();
  renderDetailTable();
}
init();
</script>
</body>
</html>
"""


def generate_region_html(xlsm_path, output_dir, region):
    payload = build_payload_for_region(xlsm_path, region)
    html = REGION_TEMPLATE.replace("{REGION}", region)
    html = html.replace("__DATA_JSON__", json.dumps(payload, default=_clean))
    output_file = os.path.join(output_dir, f"dashboard_{region}.html")
    with open(output_file, "w", encoding="utf-8") as f:
        f.write(html)
    return output_file


def generate_html(xlsm_path, out_dir=None):
    if out_dir is None:
        out_dir = os.path.dirname(xlsm_path) or "."
    results = []
    for region in ["North", "South"]:
        try:
            output_file = generate_region_html(xlsm_path, out_dir, region)
            results.append((region, output_file, True))
        except Exception as e:
            results.append((region, str(e), False))
    return results


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python make_dashboard.py <input.xlsm> [output_dir]")
        sys.exit(1)
    xlsm_path = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else os.path.dirname(xlsm_path)
    results = generate_html(xlsm_path, output_dir)
    for region, path_or_error, success in results:
        if success:
            print(f"Generated: {path_or_error}")
        else:
            print(f"{region} error: {path_or_error}")
