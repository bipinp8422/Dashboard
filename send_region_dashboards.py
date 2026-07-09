#!/usr/bin/env python3
"""
send_region_dashboards.py
==========================

Extends make_dashboard.py to:
  1. Build the full dashboard dataset from a raw .xlsm export
  2. Generate a North-only and a South-only variant (each opens straight on
     that region's tab)
  3. Optionally email each variant to the right recipient list as an
     attachment

USAGE
-----
    pip install pandas openpyxl

    # Just generate the two files (safe, no email sent):
    python send_region_dashboards.py report.xlsm

    # Generate AND send both emails via SMTP:
    python send_region_dashboards.py report.xlsm --send

    # Generate AND open both emails as Outlook drafts (Windows + Outlook desktop,
    # no password needed -- review and click Send yourself in Outlook):
    python send_region_dashboards.py report.xlsm --send --via outlook

    # Same, but send immediately without opening a draft:
    python send_region_dashboards.py report.xlsm --send --via outlook --outlook-send-immediately

Sending mail requires SMTP settings. For CLI use, set these as environment
variables (nothing is hardcoded or stored in this file):

    SMTP_HOST       e.g. smtp.office365.com
    SMTP_PORT       e.g. 587
    SMTP_USER       your full email address
    SMTP_PASSWORD   your mailbox password or app password
    SMTP_FROM       (optional) defaults to SMTP_USER

Example (Windows PowerShell):
    $env:SMTP_HOST="smtp.office365.com"
    $env:SMTP_PORT="587"
    $env:SMTP_USER="you@canon.co.in"
    $env:SMTP_PASSWORD="your-app-password"
    python send_region_dashboards.py report.xlsm --send

Example (Mac/Linux):
    export SMTP_HOST=smtp.office365.com
    export SMTP_PORT=587
    export SMTP_USER=you@canon.co.in
    export SMTP_PASSWORD=your-app-password
    python send_region_dashboards.py report.xlsm --send

send_email() also accepts an explicit `smtp_config` dict (host/port/user/
password/sender) instead of environment variables -- this is what the
Streamlit app (streamlit_app.py) uses so a "Send Now" button can fire an
email with the attachment included automatically, without needing shell
environment variables set. CLI behaviour is unchanged: no smtp_config ->
falls back to the SMTP_* environment variables exactly as before.

make_dashboard.py must sit in the same folder as this script -- it's
imported directly for the data-processing and HTML-rendering functions.
"""

import argparse
import mimetypes
import os
import smtplib
import ssl
import subprocess
import sys
import tempfile
from email.message import EmailMessage
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook as load_xlsx_workbook

from make_dashboard import load_workbook, build_dataset, render_html, filter_by_region

# Optional convenience: if a .env file sits next to this script (see
# .env.example), load SMTP_* values from it automatically so you don't have
# to `export` them by hand every time. Silently does nothing if the
# python-dotenv package isn't installed or no .env file is present -- CLI
# behaviour with real environment variables is unaffected either way.
try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).parent / ".env")
except ImportError:
    pass

# ---------------------------------------------------------------------------
# Recipient lists -- edit these if the distribution list changes.
# ---------------------------------------------------------------------------
NORTH_TO = [
    "rajender.singh@canon.co.in", "Vijay.Verma@canon.co.in", "rishabh.kapoor@canon.co.in",
    "shireesh.vaishnav@canon.co.in", "amit.nanchahal@canon.co.in", "abhishek.singh026@canon.co.in",
    "Satender.Yadav@canon.co.in", "Jaideep.SinghTEMP@canon.co.in",
]
NORTH_CC = [
    "KSHarihara.Prasad@canon.co.in", "Garima.Mohendroo@canon.co.in", "Ujjwal.JoshiTEMP@canon.co.in",
    "shashanks@denave.com", "ce-nikhil.verma@denave.com", "yunus.khan@canon.co.in",
]

SOUTH_TO = [
    "Mudari.satyanarayana@canon.co.in", "Al.Subramanian@canon.co.in", "sarath.nath@canon.co.in",
    "asiva.kumar@canon.co.in", "praveen.k@canon.co.in", "shivaraj.shettyk@canon.co.in",
    "Harikrishna.C@canon.co.in", "Shankar.d@canon.co.in",
]
SOUTH_CC = [
    "KSHarihara.Prasad@canon.co.in", "Garima.Mohendroo@canon.co.in", "sujesh.soman@canon.co.in",
    "Ujjwal.JoshiTEMP@canon.co.in", "shashanks@denave.com", "ce-nikhil.verma@denave.com",
]

# The only sheets kept in the region-filtered Excel attachment -- any other
# sheet present in the source workbook (e.g. helper/scratch tabs) is
# dropped. Order here is also the order they end up in in the output file.
ALLOWED_SHEETS = [
    "Achievement_Summary",
    "Car Counter Count",
    "Alpha Champs",
    "Target vs Achievement",
    "Alpha Program Sell-Out",
    "Raw Data",
]


# The tab-bar markup that lets a viewer flip between All/North/South. A
# region-only dashboard has nothing else to flip to (the other region's rows
# were never loaded into it), so this block is stripped out entirely rather
# than just having its "active" tab changed.
_FILTERS_BLOCK = (
    '<div class="filters" id="regionFilters">\n'
    '    <button class="tab active" data-region="All">All Regions</button>\n'
    '    <button class="tab" data-region="North">North</button>\n'
    '    <button class="tab" data-region="South">South</button>\n'
    '  </div>'
)


def render_region_only_dashboard(raw, tgt, region: str, source_name: str) -> tuple[str, dict]:
    """Build a dashboard whose underlying dataset ONLY contains `region`'s
    rows -- not just a tab pre-selected on a dataset that still secretly
    contains the other region. Returns (html, meta)."""
    r_raw, r_tgt = filter_by_region(raw, tgt, region)
    data, meta = build_dataset(r_raw, r_tgt)
    meta["source_file"] = source_name
    html = render_html(data, meta)

    # There's only one region in the data now, so drop the switcher entirely
    # instead of leaving a tab bar that implies other views exist.
    html = html.replace(_FILTERS_BLOCK, "")
    # Make the region scope visible in the title/header text itself.
    html = html.replace(
        "Daily Performance Cockpit</title>",
        f"Daily Performance Cockpit — {region}</title>",
    )
    html = html.replace(
        "<h1>Daily Performance Cockpit</h1>",
        f"<h1>Daily Performance Cockpit <span style=\"color:var(--{region.lower()})\">— {region}</span></h1>",
    )
    return html, meta


def _detect_header_row_and_region_col_ws(ws, max_scan: int = 6):
    """Scan the first few rows of an openpyxl worksheet for the cell that
    literally reads "Region" -- that's the real header row for sheets like
    these, which often have title/threshold rows above it. Returns
    (header_row_index, region_column_index), both 1-based, or (None, None)
    if no such cell is found in the first `max_scan` rows.
    """
    for r in range(1, max_scan + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip().lower() == "region":
                return r, c
    return None, None


def _convert_to_xlsx_if_needed(source_path: Path) -> Path:
    """openpyxl can only read/write .xlsx/.xlsm, not the binary .xlsb
    format. If the source is .xlsb, convert it to .xlsx with LibreOffice
    first -- a real spreadsheet-engine conversion, so every sheet's fonts,
    fills, column widths, merged cells, and number formats come through
    unchanged. Returns the source path itself if no conversion is needed.
    """
    if source_path.suffix.lower() in (".xlsx", ".xlsm"):
        return source_path

    tmp_dir = Path(tempfile.mkdtemp(prefix="xlsb_convert_"))
    result = subprocess.run(
        ["soffice", "--headless", "--convert-to", "xlsx", "--outdir", str(tmp_dir), str(source_path)],
        capture_output=True, text=True, timeout=120,
    )
    converted = tmp_dir / (source_path.stem + ".xlsx")
    if result.returncode != 0 or not converted.exists():
        raise RuntimeError(
            f"LibreOffice conversion of {source_path.name} to .xlsx failed: "
            f"{result.stderr or result.stdout}"
        )
    return converted


def build_region_source_workbook(source_path: Path, region: str, out_path: Path) -> Path:
    """Build the region-filtered attachment by editing a real copy of the
    ORIGINAL source workbook -- same sheet names, same fonts/fills/column
    widths/merged header cells/number formats -- instead of rebuilding it
    cell-by-cell from raw pandas values (which threw all formatting away).

    Only these sheets are kept in the output, in this order -- any other
    sheet present in the source workbook is dropped entirely:
        Achievement_Summary, Car Counter Count, Alpha Champs,
        Target vs Achievement, Alpha Program Sell-Out, Raw Data

    For each kept sheet, the header row + Region column are auto-detected
    (same logic as before), and every data row whose Region value doesn't
    match `region` is deleted outright (bottom-to-top, so row indices stay
    valid). Sheets with no detectable "Region" column have their rows left
    untouched rather than guessed at.

    Handles .xlsb sources transparently by converting to .xlsx via
    LibreOffice first (openpyxl can't read/write .xlsb directly).
    """
    source_path = Path(source_path)
    out_path = Path(out_path)

    xlsx_source = _convert_to_xlsx_if_needed(source_path)
    wb = load_xlsx_workbook(xlsx_source)

    # Drop any sheet not on the allowed list.
    for sheet_name in list(wb.sheetnames):
        if sheet_name not in ALLOWED_SHEETS:
            wb.remove(wb[sheet_name])

    # Put the surviving sheets in the exact order requested.
    ordered = [name for name in ALLOWED_SHEETS if name in wb.sheetnames]
    wb._sheets.sort(key=lambda ws: ordered.index(ws.title))

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row, region_col = _detect_header_row_and_region_col_ws(ws)
        if header_row is None:
            continue

        target = region.strip().lower()
        for r in range(ws.max_row, header_row, -1):
            val = ws.cell(row=r, column=region_col).value
            val_str = str(val).strip().lower() if val is not None else ""
            # Only drop rows that explicitly belong to a DIFFERENT region.
            # Blank-region rows (Total/subtotal rows, spacer rows) are left
            # alone -- they're not data for the other region, so removing
            # them would just be silently deleting part of the original
            # sheet's own structure (e.g. a summary sheet's "Total" row).
            if val_str and val_str != target:
                ws.delete_rows(r, 1)

    wb.save(out_path)
    return out_path


def build_region_files(xlsm_path: Path, out_dir: Path, month_label: str):
    raw, tgt = load_workbook(xlsm_path)

    stem = xlsm_path.stem
    north_path = out_dir / f"{stem}_NORTH.html"
    south_path = out_dir / f"{stem}_SOUTH.html"

    north_html, meta = render_region_only_dashboard(raw, tgt, "North", xlsm_path.name)
    south_html, _ = render_region_only_dashboard(raw, tgt, "South", xlsm_path.name)

    north_path.write_text(north_html, encoding="utf-8")
    south_path.write_text(south_html, encoding="utf-8")

    return north_path, south_path, meta


def _as_path_list(attachment_paths) -> list[Path]:
    """Accept either a single path or a list/tuple of paths and always
    return a list of Path objects, so send_email / send_via_outlook can
    attach one file (e.g. just the dashboard) or several (e.g. the
    dashboard plus the region-filtered source workbook)."""
    if isinstance(attachment_paths, (list, tuple)):
        return [Path(p) for p in attachment_paths]
    return [Path(attachment_paths)]


def _guess_maintype_subtype(path: Path) -> tuple[str, str]:
    ctype, _ = mimetypes.guess_type(str(path))
    if ctype is None:
        ctype = "application/octet-stream"
    maintype, _, subtype = ctype.partition("/")
    return maintype, subtype or "octet-stream"


def send_via_outlook(to_addrs, cc_addrs, subject, body, attachment_paths, send_immediately: bool = False):
    """Send (or open) an email through the local Outlook desktop app via COM
    automation -- no SMTP password needed, since it uses whatever account
    Outlook is already signed into on this machine. Windows + Outlook
    desktop only.

    attachment_paths can be a single path or a list of paths (e.g. the HTML
    dashboard plus a region-filtered source workbook) -- all of them are
    attached to the same email.

    If send_immediately is False (the default), the email is opened as a
    normal Outlook draft with the attachment(s) already added, so you can
    look it over and click Send yourself. If True, it's sent immediately
    without opening a window.
    """
    try:
        import win32com.client
    except ImportError as e:
        raise RuntimeError(
            "This requires the 'pywin32' package (Windows + Outlook desktop only). "
            "Install it with: pip install pywin32"
        ) from e

    paths = _as_path_list(attachment_paths)
    for p in paths:
        if not p.exists():
            raise FileNotFoundError(f"Attachment not found: {p}")

    outlook = win32com.client.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0)  # olMailItem
    mail.To = "; ".join(to_addrs)
    if cc_addrs:
        mail.CC = "; ".join(cc_addrs)
    mail.Subject = subject
    mail.Body = body
    for p in paths:
        mail.Attachments.Add(str(p.resolve()))

    if send_immediately:
        mail.Send()
    else:
        mail.Display()
    return mail


def send_email(to_addrs, cc_addrs, subject, body, attachment_paths, smtp_config: dict | None = None):
    """Send an email with one or more files attached.

    attachment_paths can be a single path or a list of paths (e.g. the HTML
    dashboard plus a region-filtered source workbook) -- all of them are
    attached to the same email, each with its own correct MIME type guessed
    from its extension (so an .xlsx doesn't get mislabeled as HTML, etc).

    smtp_config, if given, is a dict with any of: host, port, user, password,
    sender. Any key left out (or the whole dict left as None) falls back to
    the SMTP_HOST / SMTP_PORT / SMTP_USER / SMTP_PASSWORD / SMTP_FROM
    environment variables, exactly as before -- so existing CLI usage
    (`--send` with env vars set) is unaffected. This lets a caller like the
    Streamlit app supply credentials entered in the UI instead of requiring
    shell environment variables.
    """
    cfg = smtp_config or {}
    host = cfg.get("host") or os.environ.get("SMTP_HOST")
    port = int(cfg.get("port") or os.environ.get("SMTP_PORT", "587"))
    user = cfg.get("user") or os.environ.get("SMTP_USER")
    password = cfg.get("password") or os.environ.get("SMTP_PASSWORD")
    sender = cfg.get("sender") or os.environ.get("SMTP_FROM") or user

    missing = [k for k, v in [("host/SMTP_HOST", host), ("user/SMTP_USER", user), ("password/SMTP_PASSWORD", password)] if not v]
    if missing:
        raise ValueError(
            "Missing required SMTP setting(s): " + ", ".join(missing) +
            ". Provide them via smtp_config, or set the matching SMTP_* environment variable before running with --send."
        )

    paths = _as_path_list(attachment_paths)
    for p in paths:
        if not p.exists():
            raise FileNotFoundError(f"Attachment not found: {p}")

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = sender
    msg["To"] = ", ".join(to_addrs)
    if cc_addrs:
        msg["Cc"] = ", ".join(cc_addrs)
    msg.set_content(body)

    for p in paths:
        maintype, subtype = _guess_maintype_subtype(p)
        with open(p, "rb") as f:
            msg.add_attachment(f.read(), maintype=maintype, subtype=subtype, filename=p.name)

    all_recipients = to_addrs + cc_addrs
    context = ssl.create_default_context()
    try:
        with smtplib.SMTP(host, port, timeout=15) as server:
            server.starttls(context=context)
            server.login(user, password)
            server.send_message(msg, from_addr=sender, to_addrs=all_recipients)
    except (TimeoutError, OSError) as e:
        raise ConnectionError(
            f"Could not reach {host}:{port} within 15 seconds ({e}). "
            "This almost always means either (a) a firewall/network is blocking outbound SMTP "
            "from this machine, or (b) for Office 365, SMTP AUTH is disabled on this mailbox by "
            "default and needs an admin to enable it (Microsoft 365 admin center -> this user -> "
            "Mail -> Manage email apps -> enable 'Authenticated SMTP'). "
            "If you already have Outlook desktop working, switching to the Outlook sending method "
            "avoids this entirely."
        ) from e

    print(f"Sent: {subject}  ->  {len(all_recipients)} recipients")


def region_email_content(region: str, month_label: str):
    """Subject/body text for a region email. Shared by the CLI (`--send`)
    and the Streamlit app so both always send identical wording."""
    to_addrs = NORTH_TO if region == "North" else SOUTH_TO
    cc_addrs = NORTH_CC if region == "North" else SOUTH_CC

    subject = f"Denave x Canon CPP - Daily Performance Dashboard ({region}) - {month_label}"
    body = (
        "Hi all,\n\n"
        f"Please find attached the daily performance dashboard for the {region} region for {month_label} "
        "(month-to-date).\n\n"
        f"The dashboard opens directly on the {region} view and includes:\n"
        "- Daily revenue trend and pacing vs target\n"
        "- Day-by-day breakdown (Day Explorer) - click any day for top reps, category mix, and top cities\n"
        "- Product category mix\n"
        "- Top and bottom performers\n"
        "- FOM-wise team rollup\n\n"
        "It's a standalone HTML file - just open it in any browser, no installation needed.\n\n"
        "Regards,"
    )
    return to_addrs, cc_addrs, subject, body


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("input", help="Path to the Sales_Representative_Target_vs_Achievement_Report_Denave_<Month>.xlsm file")
    parser.add_argument("--send", action="store_true", help="Actually send the emails. Uses SMTP_* environment variables by default. Without this flag, only the two HTML files are generated.")
    parser.add_argument("--via", choices=["smtp", "outlook"], default="smtp", help="How to send when --send is used: 'smtp' (default, needs SMTP_* env vars) or 'outlook' (Windows + Outlook desktop, no password needed, opens as a draft unless --outlook-send-immediately is also given)")
    parser.add_argument("--outlook-send-immediately", action="store_true", help="With --via outlook, send immediately instead of opening a draft for review")
    parser.add_argument("-o", "--output-dir", default=None, help="Directory to write the North/South HTML files (default: same folder as input)")
    args = parser.parse_args()


    in_path = Path(args.input)
    if not in_path.exists():
        sys.exit(f"File not found: {in_path}")

    out_dir = Path(args.output_dir) if args.output_dir else in_path.parent
    out_dir.mkdir(parents=True, exist_ok=True)

    # .xlsb needs converting once up front (LibreOffice) -- both the HTML
    # dashboard data and the Excel attachments are then built from that same
    # converted file, so we don't depend on the 'pyxlsb' package at all.
    print(f"Reading {in_path.name} ...")
    working_path = _convert_to_xlsx_if_needed(in_path)
    if working_path != in_path:
        print(f"Converted .xlsb -> {working_path.name} (via LibreOffice)")

    north_path, south_path, meta = build_region_files(working_path, out_dir, None)
    print(f"Wrote {north_path}")
    print(f"Wrote {south_path}")

    stem = in_path.stem
    north_xlsx = out_dir / f"{stem}_NORTH.xlsx"
    south_xlsx = out_dir / f"{stem}_SOUTH.xlsx"
    build_region_source_workbook(working_path, "North", north_xlsx)
    build_region_source_workbook(working_path, "South", south_xlsx)
    print(f"Wrote {north_xlsx}")
    print(f"Wrote {south_xlsx}")

    if not args.send:
        print("\n--send not passed, so no email was sent. Attach these files manually, or re-run with --send.")
        return

    month_label = meta["month_label"]

    north_to, north_cc, north_subject, north_body = region_email_content("North", month_label)
    south_to, south_cc, south_subject, south_body = region_email_content("South", month_label)

    north_attachments = [north_path, north_xlsx]
    south_attachments = [south_path, south_xlsx]

    if args.via == "outlook":
        send_via_outlook(north_to, north_cc, north_subject, north_body, north_attachments, send_immediately=args.outlook_send_immediately)
        send_via_outlook(south_to, south_cc, south_subject, south_body, south_attachments, send_immediately=args.outlook_send_immediately)
    else:
        send_email(north_to, north_cc, north_subject, north_body, north_attachments)
        send_email(south_to, south_cc, south_subject, south_body, south_attachments)


if __name__ == "__main__":
    main()
